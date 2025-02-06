import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import io
import os
from datetime import datetime
import zipfile

# =====================================================
# =========== 1. Definición de funciones ==============
# =====================================================


def clean_and_format_excel(input_file, master_file, output_file):
    # Carga el reporte de pedidos
    df = pd.read_excel(input_file, dtype={"Id del pedido": str})

    # Selecciona columnas relevantes y renombra
    df_cleaned = df[
        [
            "Fecha de creación",  # A
            "Id del pedido",  # F
            "SKU de la oferta",  # Q
            "Cantidad",  # U
            "Largo del paquete (s)",  # BM
            "Alto del paquete (s)",  # BN
            "Ancho del paquete (s)",  # BO
            "Peso del paquete (s)",  # BP
        ]
    ].copy()

    # Inserta columna 'Número' (conteo simple de filas)
    df_cleaned.insert(4, "Número", range(1, len(df_cleaned) + 1))

    # Limpia contenido de las columnas de dimensiones
    df_cleaned["Largo del paquete (s)"] = ""
    df_cleaned["Alto del paquete (s)"] = ""
    df_cleaned["Ancho del paquete (s)"] = ""
    df_cleaned["Peso del paquete (s)"] = ""

    # Renombra columnas
    df_cleaned.columns = [
        "Fecha de creación",
        "Id del pedido",
        "SKU de la oferta",
        "Cantidad",
        "Número",
        "Largo del paquete (s)",
        "Alto del paquete (s)",
        "Ancho del paquete (s)",
        "Peso del paquete (s)",
    ]

    # Carga el archivo maestro
    master_df = pd.read_excel(master_file, usecols=[2], skiprows=1)
    master_skus = master_df["MODELO"].dropna().astype(str).str.strip().tolist()

    # Extrae SKU real de la columna "SKU de la oferta"
    def extract_sku(sku_offer):
        if not isinstance(sku_offer, str):
            return None

        # Sort master SKUs by length (longest first) to prevent partial matches
        sorted_skus = sorted(master_skus, key=len, reverse=True)

        for sku in sorted_skus:
            if sku_offer.lower().startswith(sku.lower()):
                return sku  # Return the first (longest) match found

        return None

    # Inserta la nueva columna "SKU"
    df_cleaned.insert(
        df_cleaned.columns.get_loc("SKU de la oferta") + 1,
        "SKU",
        df_cleaned["SKU de la oferta"].apply(extract_sku),
    )

    # Guarda a un archivo Excel en disco (para el siguiente paso)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_cleaned.to_excel(writer, index=False, sheet_name="Sheet1")
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        # Formatos
        black_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        white_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Encabezados en negro
        for col in worksheet.iter_cols(
            min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column
        ):
            for cell in col:
                cell.fill = black_fill
                cell.font = white_font
                cell.alignment = center_alignment

        # Alineación y bordes en el resto
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border


def update_working_file(working_file_path, master_file_path, output_file_path):
    # Lee el archivo de trabajo y el maestro
    working_df = pd.read_excel(working_file_path, dtype={"Id del pedido": str})
    master_df = pd.read_excel(master_file_path, skiprows=1)

    # Limpia nombre de columnas
    working_df.columns = working_df.columns.str.strip()
    master_df.columns = master_df.columns.str.strip()

    # Estandariza SKU a minúsculas
    working_df["SKU"] = working_df["SKU"].astype(str).str.strip().str.lower()
    master_df["MODELO"] = master_df["MODELO"].astype(str).str.strip().str.lower()

    unmatched_skus = []

    # Itera sobre cada SKU en el archivo de trabajo
    for index, row in working_df.iterrows():
        sku = row["SKU"]
        cantidad = row["Cantidad"]

        # Busca el SKU en el archivo maestro
        matched_row = master_df[master_df["MODELO"] == sku]

        if not matched_row.empty:
            alto = matched_row.iloc[0]["ANCHO (cm)"]  # Alto = ANCHO (cm)
            ancho = matched_row.iloc[0]["ALTO (cm)"]  # Ancho = ALTO (cm)
            largo = matched_row.iloc[0]["LARGO (cm)"]

            # Actualiza columnas (multiplicamos "alto" por la cantidad)
            working_df.at[index, "Alto del paquete (s)"] = alto * cantidad
            working_df.at[index, "Largo del paquete (s)"] = largo
            working_df.at[index, "Ancho del paquete (s)"] = ancho
        else:
            unmatched_skus.append(sku)

    # Guarda cambios
    working_df.to_excel(output_file_path, index=False)

    # Mensajes de debug
    if unmatched_skus:
        print(
            f"Los siguientes SKUs no se encontraron en el archivo maestro: {set(unmatched_skus)}"
        )
    else:
        print("Todos los SKUs se encontraron y se procesaron correctamente.")

    print(f"Archivo de trabajo actualizado y guardado en {output_file_path}")


def process_orders(file_path, output_path):
    df = pd.read_excel(file_path, dtype={"Id del pedido": str})
    df.columns = df.columns.str.strip()

    # Mover "Peso del paquete (s)" para que quede al final
    df.insert(len(df.columns), "Peso del paquete (s) Temp", df["Peso del paquete (s)"])
    df = df.drop(columns=["Peso del paquete (s)"])
    df.rename(
        columns={"Peso del paquete (s) Temp": "Peso del paquete (s)"}, inplace=True
    )

    # Insertar columnas vacías para Largo Total, Alto Total, Ancho Total
    df.insert(9, "Largo Total", "")
    df.insert(10, "Alto Total", "")
    df.insert(11, "Ancho Total", "")

    grouped = df.groupby("Id del pedido")
    for order_id, group in grouped:
        if len(group) > 1:
            max_largo = group["Largo del paquete (s)"].max()
            sum_alto = group["Alto del paquete (s)"].sum()
            max_ancho = group["Ancho del paquete (s)"].max()

            df.loc[group.index[0], "Largo Total"] = max_largo
            df.loc[group.index[0], "Alto Total"] = sum_alto
            df.loc[group.index[0], "Ancho Total"] = max_ancho

            df.loc[group.index[1:], "Largo Total"] = ""
            df.loc[group.index[1:], "Alto Total"] = ""
            df.loc[group.index[1:], "Ancho Total"] = ""
        else:
            max_largo = group["Largo del paquete (s)"].iloc[0]
            sum_alto = group["Alto del paquete (s)"].iloc[0]
            max_ancho = group["Ancho del paquete (s)"].iloc[0]

            df.loc[group.index[0], "Largo Total"] = max_largo
            df.loc[group.index[0], "Alto Total"] = sum_alto
            df.loc[group.index[0], "Ancho Total"] = max_ancho

    df.to_excel(output_path, index=False)
    print(f"Archivo procesado y guardado en {output_path}")


def calculate_package_weight(file_path, output_path):
    df = pd.read_excel(file_path, dtype={"Id del pedido": str})
    df.columns = df.columns.str.strip()

    # Mover "Peso del paquete (s)" para que quede al final
    df.insert(len(df.columns), "Peso del paquete (s) Temp", df["Peso del paquete (s)"])
    df = df.drop(columns=["Peso del paquete (s)"])
    df.rename(
        columns={"Peso del paquete (s) Temp": "Peso del paquete (s)"}, inplace=True
    )

    # Insertar columnas para peso real y peso volumétrico
    df.insert(12, "Peso real", "")
    df.insert(13, "Peso volumétrico", "")

    grouped = df.groupby("Id del pedido")
    for order_id, group in grouped:
        num_products = len(group)

        # Peso real = (número de productos * 0.600) + 0.500
        peso_real = (num_products * 0.600) + 0.500
        df.loc[group.index[0], "Peso real"] = peso_real

        # Calcular peso volumétrico si existen los totales
        if (
            (group["Largo Total"].iloc[0] != "")
            and (group["Alto Total"].iloc[0] != "")
            and (group["Ancho Total"].iloc[0] != "")
        ):
            largo_total = float(group["Largo Total"].iloc[0])
            alto_total = float(group["Alto Total"].iloc[0])
            ancho_total = float(group["Ancho Total"].iloc[0])

            peso_volumetrico = round((largo_total * alto_total * ancho_total) / 5000, 1)
            df.loc[group.index[0], "Peso volumétrico"] = peso_volumetrico

        # Calcular "Peso del paquete (s)"
        if num_products == 1:
            df.loc[group.index[0], "Peso del paquete (s)"] = 1
        elif num_products in [2, 3] and peso_real < 3:
            df.loc[group.index[0], "Peso del paquete (s)"] = 3
        else:
            if not pd.isna(df.loc[group.index[0], "Peso volumétrico"]):
                peso_volumetrico_local = df.loc[group.index[0], "Peso volumétrico"]
                peso_del_paquete = int(
                    round(peso_real + 0.75 * (peso_volumetrico_local - peso_real))
                )
                df.loc[group.index[0], "Peso del paquete (s)"] = peso_del_paquete

    df.to_excel(output_path, index=False)
    print(f"Archivo procesado y guardado en {output_path}")


def generate_shipping_labels(final_working_file, output_folder):
    final_df = pd.read_excel(final_working_file)
    final_df.columns = final_df.columns.str.strip()

    # Filtrar filas que tengan valores en las columnas indicadas
    filtered_df = final_df.dropna(
        subset=[
            "Largo Total",
            "Alto Total",
            "Ancho Total",
            "Peso real",
            "Peso volumétrico",
            "Peso del paquete (s)",
        ]
    )

    # Ordenar por "Número"
    filtered_df = filtered_df.sort_values(by="Número")

    csv_files = []  # Lista de tuplas (nombre_csv, contenido_csv)
    caja_orders = []  # IDs de pedidos con "Alto Total" > 50

    # Chunk de 50 filas por archivo
    for i in range(0, len(filtered_df), 50):
        chunk = filtered_df.iloc[i : i + 50]

        # Detectar pedidos con "Alto Total" > 50
        caja_orders.extend(
            chunk.loc[chunk["Alto Total"] > 50, "Id del pedido"].astype(str).tolist()
        )

        shipping_df = pd.DataFrame(
            {
                "pedido": chunk["Id del pedido"],
                "numero_guias": 1,
                "valor_declarado": "",
                "largo_paquete": chunk["Largo Total"].astype(int),
                "alto_paquete": chunk["Alto Total"].astype(int),
                "ancho_paquete": chunk["Ancho Total"].astype(int),
                "peso_paquete": chunk["Peso del paquete (s)"].astype(int),
            }
        )

        first_num = int(chunk["Número"].iloc[0])
        last_num = int(chunk["Número"].iloc[-1])
        first_date = chunk["Fecha de creación"].iloc[0]
        last_date = chunk["Fecha de creación"].iloc[-1]

        def reformat_date(date_obj):
            return date_obj.strftime("%d-%m-%Y-%H%M%p")

        first_date_str = reformat_date(
            pd.to_datetime(first_date, format="%d/%m/%Y - %I:%M %p")
        )
        last_date_str = reformat_date(
            pd.to_datetime(last_date, format="%d/%m/%Y - %I:%M %p")
        )

        # Nombre de archivo CSV
        file_name = f"{first_num}_{first_date_str} - {last_num}_{last_date_str}.csv"

        # Convertir a CSV en memoria
        csv_buffer = io.StringIO()
        shipping_df.to_csv(csv_buffer, index=False)
        csv_content = csv_buffer.getvalue()

        # Agregar a la lista
        csv_files.append((file_name, csv_content))

    # Crear archivo TXT si hay pedidos con "Alto Total" > 50
    caja_txt_content = None
    if caja_orders:
        caja_txt_content = "Órdenes que requieren caja (Alto Total > 50):\n\n"
        caja_txt_content += "\n".join(caja_orders)

    return csv_files, caja_txt_content


# =====================================================
# ======== 2. Lógica principal de la app ==============
# =====================================================


def main():
    st.title("Generador de archivos de Envío - Olimba")

    st.markdown(
        """
    **Instrucciones**  
    1. (Opcional) Descarga el archivo **Master Medidas** si deseas revisarlo.  
    2. (Opcional) Carga una versión actualizada de **Master Medidas** si lo necesitas (esto **eliminará** el Master actual).  
    3. **Sube** el archivo **Reporte de Pedidos** (obligatorio).  
    4. Presiona **Generar Archivos** para obtener el ZIP con el Excel final y los CSV de guías.  
    """
    )

    # 1) Descarga del archivo Master Medidas si existe
    st.subheader("1. Descargar archivo Master Medidas (opcional)")
    master_file_path = "2-Master Medidas.xlsx"
    if os.path.exists(master_file_path):
        with open(master_file_path, "rb") as f:
            st.download_button(
                label="Descargar Master Medidas actual",
                data=f,
                file_name="Master Medidas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # 2) Subir un nuevo archivo Master Medidas (reemplaza el existente)
    st.subheader("2. Subir archivo Master Medidas actualizado (opcional)")
    uploaded_master_file = st.file_uploader(
        label="Sube un nuevo archivo Master Medidas para reemplazar el actual",
        type=["xlsx"],
    )

    if uploaded_master_file is not None:
        # Eliminar el archivo original si existe
        if os.path.exists(master_file_path):
            os.remove(master_file_path)

        # Guardar el nuevo archivo con el nombre estándar
        with open(master_file_path, "wb") as f:
            f.write(uploaded_master_file.read())

        st.success("Se ha reemplazado el archivo Master Medidas con éxito.")

    # 3) Subir Reporte de Pedidos (obligatorio)
    st.subheader("3. Subir el archivo 'Reporte de Pedidos' (obligatorio)")
    uploaded_reporte_file = st.file_uploader(
        label="Sube tu archivo Excel con el Reporte de Pedidos", type=["xlsx"]
    )

    # Buffer para el ZIP final
    zip_buffer = io.BytesIO()

    # Botón para generar
    if st.button("Generar Archivos"):
        if not uploaded_reporte_file:
            st.error("Debes subir un archivo 'Reporte de Pedidos' para continuar.")
            return

        # -----------------------------------------------------
        # Pipeline de generación en memoria
        # -----------------------------------------------------
        file_1 = "1_archivo_de_trabajo.xlsx"
        file_2 = "2_archivo_de_trabajo.xlsx"
        file_3 = "3_archivo_de_trabajo.xlsx"
        file_4 = "4_archivo_de_trabajo.xlsx"

        # Guardar el archivo "Reporte de Pedidos" en disco temporal
        input_file_path = "temp_reporte_pedidos.xlsx"
        with open(input_file_path, "wb") as f:
            f.write(uploaded_reporte_file.read())

        # 1) clean_and_format_excel
        clean_and_format_excel(
            input_file=input_file_path, master_file=master_file_path, output_file=file_1
        )

        # 2) update_working_file
        update_working_file(
            working_file_path=file_1,
            master_file_path=master_file_path,
            output_file_path=file_2,
        )

        # 3) process_orders
        process_orders(file_path=file_2, output_path=file_3)

        # 4) calculate_package_weight
        calculate_package_weight(file_path=file_3, output_path=file_4)

        # 5) generate_shipping_labels (devuelve CSVs y posible TXT)
        csv_files, caja_txt_content = generate_shipping_labels(
            final_working_file=file_4,
            output_folder="temp_folder_guias",  # No se usa, pero se mantiene por compatibilidad
        )

        # Nombre final para el Excel de salida
        timestamp_str = datetime.now().strftime("%d%m%Y_%H%M")
        final_excel_name = f"archivo_de_trabajo_{timestamp_str}.xlsx"

        # Crear ZIP en memoria
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            # Agregar el archivo Excel final
            zf.write(file_4, arcname=final_excel_name)

            # Agregar CSVs en carpeta "csv_guias"
            for csv_name, csv_content in csv_files:
                zf.writestr(f"csv_guias/{csv_name}", csv_content)

            # Agregar archivo TXT solo si hay pedidos con Alto Total > 50
            if caja_txt_content:
                zf.writestr("ordenes_con_cajas.txt", caja_txt_content)

        # Regresar el cursor del buffer a 0
        zip_buffer.seek(0)

        # Descargar ZIP
        st.success("¡Archivos generados con éxito!")
        st.download_button(
            label="Descargar ZIP con archivos",
            data=zip_buffer,
            file_name=f"archivos_envio_{timestamp_str}.zip",
            mime="application/zip",
        )

        # Limpieza de archivos temporales (opcional)
        try:
            os.remove(input_file_path)
            os.remove(file_1)
            os.remove(file_2)
            os.remove(file_3)
            os.remove(file_4)
        except:
            pass


if __name__ == "__main__":
    main()
